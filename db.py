import sqlite3
from datetime import datetime
from typing import List, Union, Dict, Any, Tuple, Optional
from aiogram.types import Message
import logging
import markups
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


class Database:
    TABLE_DEFINITIONS = {
        "users": [
            ("user_id", "TEXT PRIMARY KEY"),
            ("source", "TEXT"),
            ("first_name", "TEXT"),
            ("last_name", "TEXT"),
            ("username", "TEXT"),
            ("registration_date", "TEXT"),
            ("progress", "INTEGER DEFAULT 1"),
            ("last_activity", "TEXT"),
            ("is_active", "INTEGER DEFAULT 1"),
            ("start_message_id", "INTEGER"),
            # ("start_message_text", "TEXT")
        ],
        "answers": [
            ("answer_id", "INTEGER PRIMARY KEY AUTOINCREMENT"),
            ("user_id", "INTEGER"),
            ("question_id", "INTEGER"),
            ("answer_text", "TEXT"),
            ("answer_date", "TEXT"),
        ]
    }

    QUESTIONS = {
        1: "Загранпаспорт, действующий более 2х лет",
        2: "Действующая шенгенская виза",
        3: "Срок окончания визы",
        4: "Страна местонахождения"
    }

    def __init__(self, db_file: str) -> None:
        self.db_file = db_file
        self.connection = sqlite3.connect(db_file)
        self.connection.row_factory = sqlite3.Row
        self.cursor = self.connection.cursor()

    def initialize_database(self):
        with self.connection:
            for table_name, table_columns in self.TABLE_DEFINITIONS.items():
                if table_name != "questions":
                    self.create_table(table_name, table_columns)
        logging.info("Database initialized successfully.")

    def create_table(self, table_name: str, columns: list):
        columns_with_types = ", ".join([" ".join(column) for column in columns])
        self.cursor.execute(f"CREATE TABLE IF NOT EXISTS {table_name} ({columns_with_types})")

    def get_row_as_dict(self,
                        conditions: Dict[str, Any],
                        table_names: Union[List[str], str] = None) -> Union[Dict[str, Any], None]:
        data = {}

        if table_names is None:
            table_names = ['users', 'files']

        if isinstance(table_names, str):
            table_names = [table_names]

        for table_name in table_names:
            # Формируем строку для WHERE
            where_str = " AND ".join([f"{col} = ?" for col in conditions.keys()])

            query = f"""
                    SELECT * 
                    FROM {table_name}
                    WHERE {where_str}
                """

            row = self.cursor.execute(query, tuple(conditions.values())).fetchone()

            if row is not None:
                data.update(dict(row))

        return data if data else None

    def insert_into_table(self, table_name: str, values: Dict[str, Any]) -> None:
        with self.connection:
            cursor = self.connection.cursor()

            # Формируем строку для колонок и значения
            columns_str = ", ".join(values.keys())
            values_str = ", ".join(["?" for _ in values])

            # Формируем SQL-запрос
            sql = f"INSERT INTO {table_name} ({columns_str}) VALUES ({values_str})"

            # Выполняем запрос
            cursor.execute(sql, list(values.values()))

    def update_table(self, table_name: str, values: Dict[str, Any],
                     conditions: Dict[str, Any]) -> Union[Dict[str, Any], None]:
        """
        Обновляет значения в указанной таблице и возвращает обновленную строку.

        :param table_name: Имя таблицы для обновления.
        :param values: Словарь с новыми значениями { 'column1': new_value1, 'column2': new_value2, ... }.
        :param conditions: Словарь с условиями для WHERE { 'column1': value1, 'column2': value2, ... }.
        :return: Обновленная строка или None, если строка не найдена.
        """
        with self.connection:
            # Формируем строку для SET
            set_str = ", ".join([f"{col} = ?" for col in values.keys()])

            # Формируем строку для WHERE
            where_str = " AND ".join([f"{col} = ?" for col in conditions.keys()])

            # Формируем SQL-запрос
            sql = f"UPDATE {table_name} SET {set_str} WHERE {where_str}"

            # Выполняем запрос
            self.cursor.execute(sql, list(values.values()) + list(conditions.values()))

        # Получаем и возвращаем обновленную строку
        return self.get_row_as_dict(conditions, table_name)

    def create_and_update_user(self, message: Message, source: str) -> Dict[str, Any]:
        user_id = self.generate_unique_user_id(source, message.from_user.id)
        user_data = self.get_row_as_dict({'user_id': user_id}, ['users'])

        if user_data:
            updated_data = {
                'last_name': user_data['last_name'] or message.from_user.last_name or "Not specified",
                'progress': 1,
                'last_activity': self.get_current_time_formatted()}
            user_data = self.update_table("users", updated_data, {"user_id": user_id})
            return user_data
        else:
            user_data = {
                "user_id": user_id,
                "source": source,
                "first_name": message.from_user.first_name,
                "last_name": message.from_user.last_name or "Not specified",
                "username": message.from_user.username,
                "registration_date": self.get_current_time_formatted(),
                "progress": 1,
                "last_activity": self.get_current_time_formatted(),
                "is_active": 1,
                "start_message_id": None
            }

            self.insert_into_table("users", user_data)
        return user_data

    def move_user(self, user_id: str, forward=True) -> Dict[str, Any]:
        user_data = self.get_row_as_dict({'user_id': user_id}, 'users')
        if user_data:
            step = 1 if forward else -1
            user_data['progress'] += step
            self.update_user_progress(user_id, user_data['progress'])
        return user_data

    def update_user_progress(self, user_id: str, new_progress: int) -> Dict:
        return self.update_table("users", {"progress": new_progress}, {"user_id": user_id})

    def update_user_status(self, user_id: str, new_status: str):
        self.update_table("users", {"status": new_status}, {"user_id": user_id})

    def record_answer(self, user_data, answer, question_id):
        current_time = self.get_current_time_formatted()
        answer_text = markups.inline_button_texts.get(answer, answer)
        data = {
            "user_id": user_data["user_id"],
            "question_id": question_id,
            "answer_text": answer_text,
            "answer_date": current_time
        }
        self.insert_into_table("answers", data)

        update_data = {"last_activity": current_time}
        conditions = {"user_id": user_data["user_id"]}
        self.update_table("users", update_data, conditions)

    def get_next_question(self, current_question_id, answer):
        current_question = self.get_row_as_dict({'question_id': current_question_id}, 'questions')

        if current_question:
            next_question_id = None
            if answer in ['yes', 'russia'] or current_question_id in (4, 5):
                next_question_id = current_question['next_question_yes']
            elif answer == 'no' and current_question['next_question_no'] is not None:
                next_question_id = current_question['next_question_no']
            elif answer == 'other' and current_question['next_question_other'] is not None:
                next_question_id = current_question['next_question_other']

            if next_question_id is not None:
                return self.get_row_as_dict({'question_id': next_question_id}, 'questions')

        return None

    def get_last_answer(self, user_id: str):
        query = """
            SELECT * FROM answers 
            WHERE user_id = ? 
            ORDER BY answer_date DESC 
            LIMIT 1
        """
        row = self.cursor.execute(query, (user_id,)).fetchone()
        return dict(row) if row else None

    @staticmethod
    def get_current_time_formatted():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def generate_unique_user_id(source: str, user_id: int) -> str:
        return f"{source}_{user_id}"

    def get_user_by_username(self, username: str):
        query = "SELECT * FROM users WHERE username = ?"
        row = self.cursor.execute(query, (username,)).fetchone()
        return dict(row) if row else None

    def get_user_info_for_group_chat(self, user_id: str):
        # Получение информации о пользователе
        user_data = self.get_row_as_dict({'user_id': user_id}, ['users'])

        if user_data:
            # Формирование сообщения с информацией о пользователе
            user_info_msg = f"Информация о пользователе:\n" \
                            f"Имя: {user_data['first_name']}\n" \
                            f"Фамилия: {user_data['last_name']}\n" \
                            f"Мессенджер: {user_data['source']}\n" \
                            f"Юзернейм: @{user_data['username']}\n" \
                            f"Дата регистрации: {user_data['registration_date']}\n" \
                            f"Дата последней активности: {user_data['last_activity']}\n\n"

            answers = self.get_answers_by_user_id(user_id)
            additional_questions = []
            last_answers = {}

            for answer in answers:
                if answer['question_id'] == 0:
                    additional_questions.append(answer)
                else:
                    last_answers[answer['question_id']] = answer
            # Добавление последних ответов
            if last_answers:
                user_info_msg += "Ответы на вопросы:\n"
                for question_id, answer in last_answers.items():
                    question_text = self.QUESTIONS.get(question_id)
                    user_info_msg += f"{question_id}. {question_text}: {answer['answer_text']}\n"

            # Добавление дополнительных вопросов
            if additional_questions:
                user_info_msg += "\nДополнительные вопросы:\n"
                for answer in additional_questions:
                    user_info_msg += f"- {answer['answer_text']}\n"

            return user_info_msg

    def get_answers_by_user_id(self, user_id: str):
        return self.cursor.execute(
            "SELECT * FROM answers WHERE user_id = ?", (user_id,)
        ).fetchall()

    def get_question_text(self, question_id: int):
        question = self.get_row_as_dict({'question_id': question_id}, 'questions')
        return question['question_text'] if question else "Неизвестный вопрос"

    def get_report(self) -> str:
        # Получить общее количество пользователей
        total_users = self.cursor.execute("SELECT COUNT(*) FROM users").fetchone()[0]

        # Получить количество пользователей по источнику
        user_sources = self.cursor.execute(
            "SELECT source, COUNT(*) FROM users GROUP BY source"
        ).fetchall()
        sources_text = "\n".join([f"{source}: {count}" for source, count in user_sources])

        # Получить количество пользователей, которые не закончили опрос
        users_incomplete = self.cursor.execute(
            "SELECT COUNT(*) FROM users WHERE progress > 0"
        ).fetchone()[0]

        # Получить количество активных и заблокированных пользователей
        active_users = self.cursor.execute("SELECT COUNT(*) FROM users WHERE is_active = 1").fetchone()[0]
        blocked_users = total_users - active_users

        report = f"📊 Отчет\n\n"
        report += f"Всего пользователей: {total_users}\n"
        report += f"{sources_text}\n\n"
        report += f"Не закончивших опрос: {users_incomplete}\n\n"
        report += f"Активных: {active_users}\n"
        report += f"Заблокировано: {blocked_users}"

        return report

    def create_excel_report(self) -> Tuple[io.BytesIO, str]:
        table_names = [row[0] for row in
                       self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()
                       if row[0] != 'sqlite_sequence']

        output = io.BytesIO()
        dfs = {}
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for table_name in table_names:
                query = f"SELECT * FROM {table_name}"
                df = pd.read_sql(query, self.connection)
                df.to_excel(writer, sheet_name=table_name, index=False)
                dfs[table_name] = df

        output.seek(0)
        wb = load_workbook(output)

        for table_name in table_names:
            self.apply_table_styles(wb[table_name], dfs[table_name], 1, [20] * dfs[table_name].shape[1])
        output.seek(0)
        wb.save(output)
        output.seek(0)

        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f"{current_time}Bot_report.xlsx"

        return output, file_name

    def get_all_users_id(self, messenger: str) -> List[str]:
        """Возвращает список ID пользователей для заданного мессенджера."""
        query = "SELECT user_id FROM users WHERE source = ?"
        result = self.cursor.execute(query, (messenger,)).fetchall()
        return [row['user_id'] for row in result]

    @staticmethod
    def apply_table_styles(sheet: Worksheet,
                           dataframe: pd.DataFrame,
                           start_row: int,
                           list_with_widths: Optional[List[int]] = None,
                           table_border: str = 'thin',
                           header_border: str = 'thin',
                           table_alignment: Optional[Union[str, List[str]]] = None,
                           header_alignment: Optional[str] = None,
                           wrap_text_table: bool = False,
                           wrap_text_header: bool = False,
                           cell_colors: Optional[Dict[str, List[str]]] = None,
                           range_colors: Optional[Dict[str, List[str]]] = None,
                           condition: Optional[callable] = None) -> None:
        # Set column widths
        if list_with_widths:
            for i, width in enumerate(list_with_widths, start=1):
                sheet.column_dimensions[chr(64 + i)].width = width

        # Set borders, alignment to table
        table_range = sheet[f"A{start_row + 1}:{chr(65 + len(dataframe.columns) - 1)}{start_row + len(dataframe)}"]
        table_border_style = Border(top=Side(border_style=table_border),
                                    right=Side(border_style=table_border),
                                    bottom=Side(border_style=table_border),
                                    left=Side(border_style=table_border))
        if table_alignment:
            if isinstance(table_alignment, str):
                table_alignment_style = [
                    Alignment(horizontal=table_alignment, vertical='center', wrap_text=wrap_text_table)
                    for _ in
                    range(dataframe.shape[1])]
            else:
                table_alignment_style = [Alignment(horizontal=t, vertical='center', wrap_text=wrap_text_table) for t in
                                         table_alignment]
        else:
            table_alignment_style = []  # Пустой список, если table_alignment не определено

        for row in table_range:
            for i, cell in enumerate(row):
                cell.border = table_border_style
                if table_alignment and i < len(table_alignment_style):
                    cell.alignment = table_alignment_style[i]

        # Set borders, alignment to header
        header_range = sheet[f"A{start_row}:{chr(65 + len(dataframe.columns) - 1)}{start_row}"]
        header_border_style = Border(
            top=Side(border_style=header_border),
            right=Side(border_style=header_border),
            bottom=Side(border_style=header_border),
            left=Side(border_style=header_border)
        )
        header_alignment_style = Alignment(horizontal=header_alignment, vertical='center', wrap_text=wrap_text_header)

        for row in header_range:
            for cell in row:
                cell.border = header_border_style
                if header_alignment:
                    cell.alignment = header_alignment_style

        if cell_colors and condition:
            for cell_coordinate, colors in cell_colors.items():
                cell = sheet[cell_coordinate]
                if condition(cell.value):
                    cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=colors[1], end_color=colors[1], fill_type="solid")
        elif cell_colors:
            for cell_coordinate, colors in cell_colors.items():
                cell = sheet[cell_coordinate]
                cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
        if range_colors and condition:
            for cell_range, colors in range_colors.items():
                cells = sheet[cell_range]
                for row in cells:
                    for cell in row:
                        if condition(cell.value):
                            cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color=colors[1], end_color=colors[1], fill_type="solid")
        elif range_colors:
            for cell_range, colors in range_colors.items():
                cells = sheet[cell_range]
                for row in cells:
                    for cell in row:
                        cell.fill = PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")

    def close(self):
        if self.connection:
            self.connection.close()
            print("Database connection closed.")
